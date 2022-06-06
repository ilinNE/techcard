import openpyxl as exel
from openpyxl.styles import Border, Font, NamedStyle, Side

from .models import Product, TechCard


def create_techcard(user, techcard_form, ingridient_formset, is_semifabricate):
    new_techcard = techcard_form.save(commit=False)
    new_techcard.owner = user
    new_techcard.weight = 0
    new_techcard.price = 0
    new_techcard.save()
    for ingridient_form in ingridient_formset:
        if ingridient_form in ingridient_formset.deleted_forms:
            continue
        product, ammount, cold_waste, hot_waste = tuple(
            ingridient_form.cleaned_data.values()
        )[:4]
        weigth = ammount * product.unit_weight * cold_waste * hot_waste
        price = ammount * product.price
        new_ingridient = ingridient_form.save(commit=False)
        new_ingridient.price = price
        new_ingridient.techcard = new_techcard
        new_ingridient.save()
        new_techcard.price += price
        new_techcard.weight += weigth
    if is_semifabricate:
        new_techcard.is_semifabricate = True
        semifabricate_name = new_techcard.name + " п/ф"
        semifabricate_price = new_techcard.price / new_techcard.weight
        new_techcard.semifabricate = Product.objects.create(
            name=semifabricate_name,
            owner=user,
            price=semifabricate_price,
            is_semifabricate=True,
        )
    new_techcard.save()


def edit_techcard(
    techcard_id, techcard_form, ingridient_formset, is_semifabricate
):
    techcard = TechCard.objects.get(id=techcard_id)
    user = techcard.owner
    techcard.name = techcard_form.cleaned_data["name"]
    techcard.weight = 0
    techcard.price = 0
    techcard.ingridients.all().delete()
    for ingridient_form in ingridient_formset:
        if ingridient_form in ingridient_formset.deleted_forms:
            continue
        product = ingridient_form.cleaned_data["product"]
        weigth = ingridient_form.cleaned_data["ammount"] * product.unit_weight
        price = weigth * product.price
        new_ingridient = ingridient_form.save(commit=False)
        new_ingridient.price = price
        new_ingridient.techcard = techcard
        new_ingridient.save()
        techcard.price += price
        techcard.weight += weigth
    if is_semifabricate:
        semifabricate_name = techcard.name + " п/ф"
        semifabricate_price = techcard.price / techcard.weight
        product = techcard.semifabricate
        product.name = semifabricate_name
        product.price = semifabricate_price
        product.save()
    techcard.save()


def calculate_price(ingridient_formset):
    price = 0
    weight = 0
    for form in ingridient_formset:
        if form in ingridient_formset.deleted_forms:
            continue
        product = form.cleaned_data["product"]
        weight += form.cleaned_data["ammount"] * product.unit_weight
        price += weight * product.price
    return price / weight


def make_xlsx():
    wb = exel.Workbook()
    listname = "Первый лист"
    wb.create_sheet(title=listname, index=0)
    wb.remove(wb["Sheet"])
    ws = wb[listname]
    COLUMN_WIDTHS = {
        "A": 3,
        "B": 20,
        "C": 4,
        "D": 7,
        "E": 7,
        "F": 7,
        "G": 7,
        "H": 7,
        "I": 7,
        "J": 7,
    }
    for column, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width
    for row in range(1, 12):
        ws.row_dimensions[row].height = 20
    ws.row_dimensions[9].height = 85

    ws.merge_cells("A1:E2")
    ws.merge_cells("F1:I1")
    ws.merge_cells("F2:I2")
    ws.merge_cells("A3:I3")
    ws.merge_cells("A9:I9")
    ws.merge_cells("A10:I10")
    ws.merge_cells("A11:B11")
    ws.merge_cells("C11:D11")
    ws.merge_cells("E11:G11")
    ws.merge_cells("H11:I11")
    borderstyle = NamedStyle(name="borderstyle")
    borderstyle.font = Font(bold=True, size=20)
    bd = Side(style="thin", color="000000")
    borderstyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    cells = ws["A1":"I11"]
    for row in cells:
        for cell in row:
            cell.style = borderstyle
    return wb
